import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "히스토리 경우의수"

def fill(hex): return PatternFill("solid", fgColor=hex)
def fnt(bold=False, color="000000", size=10, italic=False, strike=False):
    return Font(bold=bold, color=color, size=size, italic=italic, strike=strike)
def aln(h="center", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def brd():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def sc(ws, row, col, val, ct="n", rf=None):
    c = ws.cell(row=row, column=col, value=val)
    c.alignment = aln(); c.border = brd()
    rf = rf or fill("FFFFFF")
    if   ct=="h1":  c.fill=fill("1F4E79"); c.font=fnt(bold=True,color="FFFFFF",size=10)
    elif ct=="h2":  c.fill=fill("2E75B6"); c.font=fnt(bold=True,color="FFFFFF",size=9)
    elif ct=="tW":  c.fill=rf; c.font=fnt(bold=True,color="CC0000",size=10)
    elif ct=="tT":  c.fill=rf; c.font=fnt(bold=True,color="CC0000",size=10)
    elif ct=="tM":  c.fill=rf; c.font=fnt(bold=True,color="0070C0",size=10)
    elif ct=="tS":  c.fill=rf; c.font=fnt(color="808080",size=10)
    elif ct=="act": c.fill=rf; c.font=fnt(size=10)
    elif ct=="cr":  c.fill=fill("EBF3FB"); c.font=fnt(bold=True,color="0070C0",size=10)
    elif ct=="bf":  c.fill=fill("FFF2CC"); c.font=fnt(color="888888",strike=True,size=10)
    elif ct=="af":  c.fill=fill("FFF2CC"); c.font=fnt(bold=True,color="CC0000",size=10)
    elif ct=="sm":  c.fill=rf; c.font=fnt(color="000000",size=10)
    elif ct=="nl":  c.fill=rf; c.font=fnt(color="BBBBBB",size=10)
    elif ct=="eq":  c.fill=rf; c.font=fnt(color="CCCCCC",italic=True,size=9)
    elif ct=="nt":  c.fill=rf; c.font=fnt(color="444444",size=9)
    elif ct=="fl":  c.fill=fill("D9E1F2"); c.font=fnt(bold=True,size=10)
    else:           c.fill=rf; c.font=fnt(size=10)
    return c

# ── 헤더 ──────────────────────────────────────────────────
H1=["#","플로우","변경자","작업","상태","상태","입찰가","입찰가","회원","회원","담당자","담당자","비고"]
H2=["","","","","변경전(취소선)","변경후(빨강굵게)","변경전","변경후","변경전","변경후","변경전","변경후",""]
ws.row_dimensions[1].height=30; ws.row_dimensions[2].height=22
for ci,h in enumerate(H1,1): sc(ws,1,ci,h,"h1")
for ci,h in enumerate(H2,1): sc(ws,2,ci,h,"h2")
for col in [1,2,3,4,13]: ws.merge_cells(start_row=1,start_column=col,end_row=2,end_column=col)

# ── 데이터 ────────────────────────────────────────────────
# (플로우, 변경자표시, 변경자타입, 작업표시,
#  stu_f, stu_t,  bid_f, bid_t,  mem_f, mem_t,  mgr_f, mgr_t,  비고)
#
# to값="★" → ITEM_CREATE 파랑굵게
# to값=None → 변경없음 (from검은+후열"=")
# from=None  → null ("-" 회색)
# to값=문자  → 변경됨 (from취소선+to빨강)

DATA=[
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ① 물건 등록
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("① 물건 등록\n[웹 직접등록]",
 "Web","tW","물건 등록",
 "미정","★",  None,None,  None,None,  None,None,
 "ITEM_CREATE / web. 관리자가 웹 폼 직접 입력. 모든 초기값 파란색굵게"),

("① 물건 등록\n[이미지 자동등록]",
 "sys","tS","물건 등록",
 "미정","★",  None,None,  None,None,  None,None,
 "ITEM_CREATE / sys. 이미지캡처 자동등록 시스템 경유. 95%+ 해당. 초기값 파란색굵게"),

("① 물건 등록\n[회원포함 등록]",
 "Web","tW","물건 등록",
 "추천","★",  None,None,  "홍길동","★",  "대표님","★",
 "ITEM_CREATE / web. 회원 포함 등록 시. 상태/회원/담당자 초기값 파란색굵게"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ② 추천 배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("② 추천 배정",
 "Web","tW","필드 변경",
 "미정","추천",  None,None,  "(없음)","홍길동",  "(없음)","대표님",
 "FIELD_CHANGE / web. stu_member+m_name+m_name_id 동시변경 → 1행 그룹화. 입찰가 null 유지"),

("② 추천 배정\n[텔레그램발송]",
 "web-telegram","tT","텔레그램\n(추천물건전달)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "TELEGRAM_SENT note=card / web-telegram. 변경없음 → 스냅샷 검은색"),

("② 추천 배정\n[chuchen자동]",
 "web-telegram","tT","필드 변경",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "FIELD_CHANGE chuchen_state: null→전달완료 / web-telegram. ★화면 컬럼 미표시★ (상태/입찰가/회원/담당자 변경없음)"),

("② 추천 배정\n[chuchen수동]",
 "Web","tW","필드 변경",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "FIELD_CHANGE chuchen_state: 신규→전달완료 / web. UI [전달완료] 버튼 클릭. ★화면 컬럼 미표시★"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ③ 입찰 확정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("③ 입찰 확정",
 "mem-telegram","tM","입찰 요청",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_BID / member-telegram. 회원 텔레그램 [입찰확정] 클릭. 변경없음"),

("③ 입찰 확정\n[자동승인]",
 "sys","tS","요청 승인\n(자동/수동)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_APPROVED / system. 자동 승인. 변경없음"),

("③ 입찰 확정\n[수동승인]",
 "Web","tW","요청 승인\n(자동/수동)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_APPROVED / web. 관리자 수동 승인. 변경없음"),

("③ 입찰 확정",
 "sys","tS","필드 변경",
 "추천","입찰",  None,None,  "홍길동",None,  "대표님",None,
 "FIELD_CHANGE stu_member: 추천→입찰 / system(자동) 또는 web(수동). 입찰가는 null 유지"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ④ 추천 취소 (미선택)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("④ 추천 취소\n(미선택)",
 "mem-telegram","tM","추천 취소 요청",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_CANCEL_CHUCHEN / member-telegram. 회원 [미선택] 클릭. 변경없음"),

("④ 추천 취소\n[자동승인]",
 "sys","tS","요청 승인\n(자동/수동)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_APPROVED / system. 변경없음"),

("④ 추천 취소\n[수동승인]",
 "Web","tW","요청 승인\n(자동/수동)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_APPROVED / web. 관리자 수동 승인. 변경없음"),

("④ 추천 취소\n(미선택)",
 "sys","tS","필드 변경",
 "추천","미정",  None,None,  "홍길동","(없음)",  "대표님","(없음)",
 "FIELD_CHANGE stu_member+m_name+m_name_id 동시변경. 입찰가 null 유지"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑤ 입찰가 전달
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑤ 입찰가 전달",
 "Web","tW","필드 변경",
 "입찰",None,  "(없음)","77,990,000",  "홍길동",None,  "대표님",None,
 "FIELD_CHANGE bidprice: null→77,990,000 / web. bid_state=전달완료 동시저장 → 자동 텔레그램 발송"),

("⑤ 입찰가 전달\n[텔레그램발송]",
 "web-telegram","tT","텔레그램\n(입찰가전달)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "TELEGRAM_SENT note=bid_price / web-telegram. 변경없음"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑥ 입찰가 확인요청 (확인 요청 전달)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑥ 확인요청 전달",
 "web-telegram","tT","텔레그램\n(확인요청전달)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "TELEGRAM_SENT note=check_request / web-telegram. 확인 버튼 포함 메시지 발송. 변경없음"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑦ 입찰가 확인
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑦ 입찰가 확인\n[텔레그램버튼]",
 "mem-telegram","tM","입찰가확인",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "PRICE_CONFIRMED note=입찰가 확인(빠른답장) / member-telegram. 텔레그램 버튼 클릭. bid_state→확인완료(화면미표시)"),

("⑦ 입찰가 확인\n[웹앱]",
 "mem-telegram","tM","입찰가확인",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "PRICE_CONFIRMED note=입찰가 확인 / member-telegram. 웹앱에서 확인 클릭. bid_state→확인완료(화면미표시)"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑧ 입찰 취소
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑧ 입찰 취소",
 "mem-telegram","tM","입찰 취소 요청",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "REQUEST_CANCEL_BID / member-telegram. 회원 텔레그램 [입찰취소] 클릭. 변경없음"),

("⑧ 입찰 취소\n[자동승인]",
 "sys","tS","요청 승인\n(자동/수동)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "REQUEST_APPROVED / system. 변경없음"),

("⑧ 입찰 취소\n[수동승인]",
 "Web","tW","요청 승인\n(자동/수동)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "REQUEST_APPROVED / web. 관리자 수동 승인. 변경없음"),

("⑧ 입찰 취소",
 "sys","tS","필드 변경",
 "입찰","미정",  "77,990,000","(없음)",  "홍길동","(없음)",  "대표님","(없음)",
 "FIELD_CHANGE stu_member+bidprice+m_name+m_name_id 동시변경"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑨ 변경/취소 안내
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑨ 변경/취소 안내",
 "Web","tW","필드 변경",
 "입찰","변경",  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "FIELD_CHANGE stu_member: 입찰→변경 / web. 관리자 직접 상태 수정"),

("⑨ 변경/취소 안내\n[텔레그램발송]",
 "web-telegram","tT","텔레그램\n(진행상태전달)",
 "변경",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "TELEGRAM_SENT note=status / web-telegram. 변경없음"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑩ 자동 만료
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑩ 자동 만료",
 "sys","tS","만료 예정 알림\n(24h 전)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "EXPIRY_NOTIFY note=24h / system. 추천 후 24h 경과 시 자동발송"),

("⑩ 자동 만료",
 "sys","tS","만료 예정 알림\n(1h 전)",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "EXPIRY_NOTIFY note=1h / system. 추천 후 47h 경과 시 자동발송"),

("⑩ 자동 만료",
 "sys","tS","자동 추천 만료",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "AUTO_EXPIRE / system. 48h 경과. AUTO_EXPIRE_ENABLED=true"),

("⑩ 자동 만료",
 "sys","tS","필드 변경",
 "추천","미정",  None,None,  "홍길동","(없음)",  "대표님","(없음)",
 "FIELD_CHANGE stu_member+m_name+m_name_id / system. 자동 초기화"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑪ 입찰일 알림
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑪ 입찰일 알림",
 "sys","tS","입찰일 알림\n(D-3)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "BID_DATE_NOTIFY note=d3 / system. BID_NOTIFY_D3=true"),

("⑪ 입찰일 알림",
 "sys","tS","입찰일 알림\n(D-2)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "BID_DATE_NOTIFY note=d2 / system"),

("⑪ 입찰일 알림",
 "sys","tS","입찰일 알림\n(D-1)",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "BID_DATE_NOTIFY note=d1 / system"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑫ 요청 반려
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑫ 요청 반려\n[입찰확정 반려]",
 "Web","tW","요청 반려",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_REJECTED / web. REQUEST_BID 반려. 상태변경 없음"),

("⑫ 요청 반려\n[추천취소 반려]",
 "Web","tW","요청 반려",
 "추천",None,  None,None,  "홍길동",None,  "대표님",None,
 "REQUEST_REJECTED / web. REQUEST_CANCEL_CHUCHEN 반려. 상태변경 없음"),

("⑫ 요청 반려\n[입찰취소 반려]",
 "Web","tW","요청 반려",
 "입찰",None,  "77,990,000",None,  "홍길동",None,  "대표님",None,
 "REQUEST_REJECTED / web. REQUEST_CANCEL_BID 반려. 상태변경 없음"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑬ 회원 재배정
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑬ 회원 재배정\n(A→B)",
 "Web","tW","필드 변경",
 "추천",None,  None,None,  "홍길동","이순신",  "대표님",None,
 "FIELD_CHANGE m_name: 홍길동→이순신 / web. stu_member/입찰가 변경없음"),

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ⑭ 물건 삭제
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
("⑭ 물건 삭제",
 "Web","tW","물건 삭제",
 "미정",None,  None,None,  "(없음)",None,  "(없음)",None,
 "ITEM_DELETE / web. 삭제 시점 스냅샷. 상태에 따라 값 다름"),
]

# ── 렌더링 ────────────────────────────────────────────────
prev_flow=None; rn=3
for idx,d in enumerate(DATA,1):
    (flow,trig,ttype,action,
     sf,st,bf,bt,mf,mt,gf,gt,note)=d
    is_new=(flow!=prev_flow); prev_flow=flow
    rf=fill("F7F9FF") if idx%2==1 else fill("FFFFFF")
    if is_new: rf=fill("EEF2FA")
    ws.row_dimensions[rn].height=40

    sc(ws,rn,1,idx,"nt",rf)

    c=ws.cell(row=rn,column=2,value=flow)
    c.fill=fill("EEF2FA") if is_new else rf
    c.font=fnt(bold=True,size=10); c.alignment=aln(); c.border=brd()

    sc(ws,rn,3,trig,ttype,rf)
    sc(ws,rn,4,action,"act",rf)

    for fv,tv,cf,ct in[(sf,st,5,6),(bf,bt,7,8),(mf,mt,9,10),(gf,gt,11,12)]:
        if fv is None and tv is None:
            sc(ws,rn,cf,"-","nl",rf); sc(ws,rn,ct,"-","nl",rf)
        elif tv=="★":
            sc(ws,rn,cf,fv,"cr",rf); sc(ws,rn,ct,"↑파랑굵게","cr",rf)
        elif tv is not None:
            sc(ws,rn,cf,fv,"bf",rf); sc(ws,rn,ct,tv,"af",rf)
        else:
            sc(ws,rn,cf,fv,"sm",rf); sc(ws,rn,ct,"=","eq",rf)

    sc(ws,rn,13,note,"nt",rf)
    rn+=1

# ── 컬럼 너비 ────────────────────────────────────────────
for col,w in{1:4,2:14,3:13,4:17,5:10,6:10,7:13,8:13,9:10,10:10,11:10,12:10,13:60}.items():
    ws.column_dimensions[get_column_letter(col)].width=w
ws.freeze_panes="E3"

# ── 범례 ─────────────────────────────────────────────────
ws2=wb.create_sheet("범례")
legs=[("표시","색/스타일","의미"),
("변경전값","회색+취소선","해당 필드 변경됨 - 이전값"),
("변경후값","빨간+굵게","해당 필드 변경됨 - 새값"),
("현재값","검은색","변경없음 - 스냅샷"),
("-","연회색","null (값 없음)"),
("↑파랑굵게","파란+굵게","ITEM_CREATE 최초등록값"),
("=","연회색이탤릭","변경없음 표시"),
("Web","빨간굵게","trigger=web"),
("web-telegram","빨간굵게","trigger=web-telegram"),
("mem-telegram","파란굵게","trigger=member-telegram"),
("sys","회색","trigger=system"),
("홍길동","일반","예시 회원명 (실제는 가변)"),
("이순신","일반","예시 회원명B (재배정 시)"),
("대표님","일반","예시 담당자명 (실제는 가변)"),
("77,990,000","일반","예시 입찰가 (실제는 가변)"),
("★화면 컬럼 미표시★","비고","chuchen_state 변경은 상태/입찰가/회원/담당자 컬럼에 안 보임"),]
for ri,row in enumerate(legs,1):
    for ci,v in enumerate(row,1):
        c=ws2.cell(row=ri,column=ci,value=v)
        c.alignment=aln(h="left"); c.border=brd()
        if ri==1: c.fill=fill("1F4E79"); c.font=fnt(bold=True,color="FFFFFF")
        elif ri%2==0: c.fill=fill("F2F2F2")
ws2.column_dimensions["A"].width=18
ws2.column_dimensions["B"].width=18
ws2.column_dimensions["C"].width=52

wb.save(r"C:\LJW\MAPS_TEST\doc\히스토리_경우의수.xlsx")
print("완료")
