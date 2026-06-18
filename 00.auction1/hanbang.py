# -*- coding: utf-8 -*-
"""
hanbang.py — 한방(karhanbang.com) 중개사무소 리스트 크롤러
- 대상: https://www.karhanbang.com/office/?topM=09  (페이지당 10개, 기본 2498페이지)
- 리스트에서: 지역 / 상호 / 대표자 / 신주소 / 사무실번호(02)
- 상세 진입해서: 부가주소 / 핸드폰번호(010)   ← 핸드폰은 리스트에 없음
- 사무실번호와 핸드폰번호는 분리 컬럼으로 저장
- 결과: xlsx 내보내기 (openpyxl)

주의: 사이트에 dotDefender WAF 가 있어 봇처럼 빠르게 연속요청하면 차단됨.
      → requests 세션(쿠키) + 브라우저 헤더 + Referer + 요청 간 지연 으로 사람처럼 천천히.
"""
from __future__ import annotations
import os, re, time, uuid, threading, json, html as _html

import requests

_DIR_HERE = os.path.dirname(os.path.abspath(__file__))
_EXPORT_DIR = os.path.join(_DIR_HERE, "hanbang_exports")

BASE = "https://www.karhanbang.com"
LIST_BASE = BASE + "/office/"
DETAIL_BASE = BASE + "/office/office_detail.asp"

_HEADERS = {
    "User-Agent": ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                   "(KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36"),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

# 컬럼 순서 (사용자 사양) — 상호 / 핸드폰 / 신주소 3개만
# - 상호 = 앞 괄호 제거 (예: (가락)학사공인중개사사무소 → 학사공인중개사사무소)
# - 핸드폰 = 상세의 010 (없으면 공란)
# - 신주소 = 상세의 소재지(풀 도로명주소). 리스트엔 동 레벨만 있어 상세에서 보강
COLUMNS = ["상호", "핸드폰", "신주소"]

# 시도(지역) 코드 — sel_sido 파라미터 (사이트 드롭다운 순서)
SIDO = [
    (1, "서울특별시"), (2, "경기도"), (3, "인천광역시"), (4, "부산광역시"),
    (5, "대구광역시"), (6, "광주광역시"), (7, "대전광역시"), (8, "울산광역시"),
    (9, "강원특별자치도"), (10, "경상남도"), (11, "경상북도"), (12, "전라남도"),
    (13, "전북특별자치도"), (14, "충청남도"), (15, "충청북도"), (16, "세종특별자치시"),
    (17, "제주특별자치도"),
]

SAVE_EVERY_PAGES = 1    # 매 페이지마다 중간저장 + 체크포인트 (크래시에도 직전 페이지까지 보호)
_CKPT_PATH = os.path.join(_EXPORT_DIR, "hanbang_checkpoint.json")
_TOTALS_CACHE = os.path.join(_EXPORT_DIR, "hanbang_region_totals.json")   # 지역별 총페이지 캐시

# ── 작업 레지스트리 (run_id -> state) ───────────────────────────
_runs = {}            # run_id -> dict
_runs_lock = threading.Lock()


# ============================================================
# 파싱 유틸
# ============================================================
def _text(s: str) -> str:
    """태그 제거 + 엔티티 복원 + 공백 정리."""
    if not s:
        return ""
    s = re.sub(r"<[^>]+>", " ", s)
    s = _html.unescape(s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def parse_list(html_text: str):
    """리스트 페이지 → [{region, name, rep, office_tel, new_addr, mem_no, param}]"""
    out = []
    m = re.search(r"<tbody>(.*?)</tbody>", html_text, re.S)
    scope = m.group(1) if m else html_text
    for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", scope, re.S):
        tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S)
        if len(tds) < 5:
            continue
        mv = re.search(r"moveDetail\('(\d+)'\s*,\s*'([^']*)'\)", tds[1])
        if not mv:
            continue
        mem_no, param = mv.group(1), mv.group(2)
        # 상호 = 앵커 텍스트에서 span(viewnum/date) 제거 후
        sub = re.sub(r"<span.*?</span>", "", tds[1], flags=re.S)
        name = _text(sub)
        # 사무실 전화 (02 등) = td3 의 tel: 링크
        mt = re.search(r"tel:([0-9\-]+)", tds[3])
        office_tel = mt.group(1) if mt else _text(tds[3])
        out.append({
            "region": _text(tds[0]),
            "name": name,
            "rep": _text(tds[2]),
            "office_tel": office_tel,
            "new_addr": _text(tds[4]),
            "mem_no": mem_no,
            "param": param,
        })
    return out


def _detail_li(html_text: str, label: str) -> str:
    """상세 <li><span>라벨</span><em>값</em></li> 에서 값 추출."""
    m = re.search(r"<li[^>]*>\s*<span[^>]*>\s*" + re.escape(label) + r"\s*</span>\s*<em[^>]*>(.*?)</em>",
                  html_text, re.S)
    return _text(m.group(1)) if m else ""


def parse_detail(html_text: str) -> dict:
    """상세 페이지 → {sojae(신주소·도로명), buga(부가주소), jibun(지번주소), mobile, office}"""
    sojae = _detail_li(html_text, "소재지")   # 도로명주소 = 신주소(풀)
    buga = _detail_li(html_text, "부가주소")
    jibun = _detail_li(html_text, "지번주소")
    office = ""
    mobile = ""
    mr = re.search(r"전화걸기\s*</span>\s*<em[^>]*>(.*?)</em>", html_text, re.S)
    if mr:
        tels = re.findall(r"tel:([0-9\-]+)", mr.group(1))
        for t in tels:
            digits = re.sub(r"[^0-9]", "", t)
            if not digits or digits == "000":
                continue
            if digits.startswith("01"):
                mobile = t            # 핸드폰 (010 등)
            elif not office:
                office = t            # 사무실 (02/지역번호)
    return {"sojae": sojae, "buga": buga, "jibun": jibun, "mobile": mobile, "office": office}


def strip_name_paren(name: str) -> str:
    """상호 앞 괄호 제거. '(가락)학사공인중개사사무소' → '학사공인중개사사무소'."""
    return re.sub(r"^\s*\([^)]*\)\s*", "", name or "").strip()


def detect_block(html_text: str) -> bool:
    return "dotDefender" in html_text or "Blocked Your Request" in html_text


# ============================================================
# 크롤 작업 (스레드)
# ============================================================
def _log(state, msg):
    with state["lock"]:
        state["lines"].append(msg)


def _now_stamp():
    return time.strftime("%Y-%m-%d %H:%M")


def _list_page_url(page: int, sido=None) -> str:
    # sido 지정 시 지역 필터, 미지정 시 사이트 기본 리스트(=서울)
    if sido:
        return f"{LIST_BASE}?topM=09&sel_sido={sido}&page={page}"
    return f"{LIST_BASE}?topM=09&page={page}"


def _detail_url(mem_no: str, param: str) -> str:
    return f"{DETAIL_BASE}?topM=09&mem_no={mem_no}&{param}"


def _detect_total_pages(html_text: str) -> int:
    pages = [int(x) for x in re.findall(r"[?&]page=(\d+)", html_text)]
    return max(pages) if pages else 0


def _fetch(session, url, referer, delay, state, tries=3):
    """지연 후 GET. dotDefender 차단 감지 시 백오프 재시도. (text, blocked)"""
    last = ""
    for attempt in range(tries):
        if state.get("cancel"):
            return "", False
        time.sleep(delay if attempt == 0 else delay * (attempt + 2))
        try:
            h = dict(_HEADERS)
            if referer:
                h["Referer"] = referer
            r = session.get(url, headers=h, timeout=20)
            r.encoding = "utf-8"
            txt = r.text
            if detect_block(txt):
                last = txt
                _log(state, f"  ⚠ WAF 차단 감지 — {(attempt+1)}/{tries} 백오프 재시도")
                continue
            return txt, False
        except Exception as e:
            last = ""
            _log(state, f"  ⚠ 요청 오류({attempt+1}/{tries}): {e}")
    return last, detect_block(last)


def _session():
    s = requests.Session()
    s.headers.update(_HEADERS)
    return s


def _warm(state, session, delay, sido=None) -> bool:
    _log(state, "세션 준비 중… (워밍)")
    warm, blocked = _fetch(session, _list_page_url(1, sido), BASE + "/", delay, state)
    if blocked or not warm:
        _log(state, "❌ 시작 실패: 첫 요청이 WAF 차단됨. 잠시 후 다시 시도하세요.")
        return False
    return True


def _crawl_one_page(state, session, page, end_label, sido, delay, rows):
    """리스트 1페이지 + 각 사무소 상세 → rows 누적."""
    state["cur_page"] = page
    list_html, blocked = _fetch(session, _list_page_url(page, sido),
                                _list_page_url(max(1, page - 1), sido), delay, state)
    if blocked or not list_html:
        _log(state, f"❌ {page}쪽 리스트 실패(차단/오류) — 건너뜀")
        return
    offices = parse_list(list_html)
    _log(state, f"[{page}/{end_label}쪽] 사무소 {len(offices)}건")
    for i, off in enumerate(offices, 1):
        if state.get("cancel"):
            return
        d_html, blocked = _fetch(session, _detail_url(off["mem_no"], off["param"]),
                                 _list_page_url(page, sido), delay, state)
        det = parse_detail(d_html) if (not blocked and d_html) else {}
        name2 = strip_name_paren(off["name"])
        mobile = det.get("mobile", "")
        sin_addr = det.get("sojae") or off["new_addr"]   # 상세 소재지(풀) 우선, 없으면 리스트 폴백
        rows.append({"상호": name2, "핸드폰": mobile, "신주소": sin_addr})
        state["count"] = state.get("count", 0) + 1
        _log(state, f"    {page}-{i} {name2} | {mobile or '없음'} | {sin_addr or '-'}")


def _run(state, start_page, end_page, delay):
    """단일/테스트 모드 — 페이지 범위 크롤 후 엑셀 1개."""
    rows = []
    try:
        os.makedirs(_EXPORT_DIR, exist_ok=True)
        session = _session()
        if not _warm(state, session, delay, None):
            state["status"] = "error"
            return
        _log(state, f"크롤 범위: {start_page}~{end_page}쪽  (지연 {delay}s)")
        for page in range(start_page, end_page + 1):
            if state.get("cancel"):
                _log(state, "⏹ 사용자 중지")
                break
            _crawl_one_page(state, session, page, end_page, None, delay, rows)
        if rows:
            path = os.path.join(_EXPORT_DIR, f"한방_{start_page}-{end_page}p_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
            _save_rows(rows, path)
            state["file"] = path
            _log(state, f"✅ 완료 — {len(rows)}건 저장: {os.path.basename(path)}")
        else:
            _log(state, "결과 0건 — 저장할 데이터 없음")
        state["status"] = "done"
    except Exception as e:
        import traceback
        _log(state, "❌ 예외: " + str(e))
        _log(state, traceback.format_exc())
        state["status"] = "error"
    finally:
        if state["status"] not in ("done", "error"):
            state["status"] = "done"


def _do_region(state, session, idx, sido, name, delay, ck, label, force=False):
    """시도 1개 크롤 — 체크포인트/재개/중간저장. 반환 'done'|'cancelled'|'skipped'|'error'."""
    key = str(idx)
    info = {} if force else ck.get(key, {})
    state["cur_region"] = name
    fpath = _region_file(idx, name)
    if info.get("done"):
        state["region_done"] = state.get("region_done", 0) + 1
        _log(state, f"⏭ {label} {name} — 이미 완료, 건너뜀 (다시 받으려면 파일 삭제 후 실행)")
        return "skipped"
    first, blocked = _fetch(session, _list_page_url(1, sido), BASE + "/", delay, state)
    if blocked or not first:
        _log(state, f"❌ {label} {name} 시작 실패(차단) — 잠시 후 재시도")
        return "error"
    total = int(info.get("total") or _detect_total_pages(first))
    rows = _load_xlsx_rows(fpath) if (os.path.exists(fpath) and not force) else []
    done_page = 0 if force else int(info.get("last_page") or 0)
    started = info.get("started_at") or _now_stamp()   # 최초 시작 시각 보존
    first_page = done_page + 1
    if done_page > 0:
        _log(state, f"↻ {label} {name} — {done_page}/{total}쪽까지 완료, 이어서 (기존 {len(rows)}건)")
    else:
        _log(state, f"▶ {label} {name} — 총 {total}쪽 시작")
    for page in range(done_page + 1, total + 1):
        if state.get("cancel"):
            break
        before = len(rows)
        _crawl_one_page(state, session, page, total, sido, delay, rows)
        if state.get("cancel"):
            del rows[before:]   # 중단된 페이지의 부분 데이터 제거(재개 시 중복 방지)
            break
        done_page = page
        if page % SAVE_EVERY_PAGES == 0:   # =1 → 매 페이지 저장 (크래시에도 직전 페이지까지 보호)
            if _safe_save(state, rows, fpath):
                ck[key] = {"done": False, "last_page": done_page, "total": total, "started_at": started}
                _save_ckpt(ck)
                if page == first_page or page % 10 == 0 or page == total:   # 로그만 솎음(저장은 매 페이지)
                    _log(state, f"  💾 저장 {done_page}/{total}쪽 ({len(rows)}건)")
            # 저장 실패(파일 잠김 등) — 체크포인트 미갱신, 데이터는 메모리 유지, 다음 저장 때 재시도하며 계속
    saved_ok = True
    if rows:
        saved_ok = _safe_save(state, rows, fpath)
        if saved_ok:
            state["file"] = fpath
            if os.path.basename(fpath) not in state["region_files"]:
                state["region_files"].append(os.path.basename(fpath))
    if state.get("cancel"):
        if saved_ok:
            ck[key] = {"done": False, "last_page": done_page, "total": total, "started_at": started}
            _save_ckpt(ck)
        _log(state, f"⏹ {name} 중단 — {done_page}/{total}쪽까지 저장({len(rows)}건). 다음 실행 때 이어서.")
        return "cancelled"
    if not saved_ok:
        _log(state, f"⚠ {name} 최종 저장 실패(파일 잠김) — 완료 표시 보류. Excel 닫고 다시 누르면 이어서 재개.")
        return "error"
    ck[key] = {"done": True, "last_page": total, "total": total, "started_at": started, "done_at": _now_stamp()}
    _save_ckpt(ck)
    state["region_done"] = state.get("region_done", 0) + 1
    _log(state, f"✅ {label} {name} 완료 — {len(rows)}건: {os.path.basename(fpath)}")
    return "done"


def _run_regions(state, delay):
    """지역별 자동 모드 — 17개 시도 순회. 시도마다 엑셀 1개. 완료 지역은 재실행 시 건너뜀."""
    try:
        os.makedirs(_EXPORT_DIR, exist_ok=True)
        session = _session()
        if not _warm(state, session, delay, None):
            state["status"] = "error"
            return
        state["region_total"] = len(SIDO)
        _log(state, f"🌐 전국 지역별 자동 크롤 — {len(SIDO)}개 시도 (지연 {delay}s · {SAVE_EVERY_PAGES}쪽마다 중간저장)")
        ck = _load_ckpt()
        for idx, (sido, name) in enumerate(SIDO, 1):
            if state.get("cancel"):
                _log(state, "⏹ 사용자 중지")
                break
            res = _do_region(state, session, idx, sido, name, delay, ck, f"[{idx}/{len(SIDO)}]")
            if res == "cancelled":
                break
        state["status"] = "done"
        _log(state, f"🏁 종료 — 완료 {state.get('region_done', 0)}/{len(SIDO)} 지역")
    except Exception as e:
        import traceback
        _log(state, "❌ 예외: " + str(e))
        _log(state, traceback.format_exc())
        state["status"] = "error"
    finally:
        if state["status"] not in ("done", "error"):
            state["status"] = "done"


def _run_one_region(state, sido_code, delay):
    """특정 시도 1개만 크롤. 사용자가 지역 선택 후 실행."""
    try:
        os.makedirs(_EXPORT_DIR, exist_ok=True)
        match = [(i, s, n) for i, (s, n) in enumerate(SIDO, 1) if s == int(sido_code)]
        if not match:
            _log(state, f"❌ 알 수 없는 지역 코드: {sido_code}")
            state["status"] = "error"
            return
        idx, sido, name = match[0]
        state["region_total"] = 1
        session = _session()
        if not _warm(state, session, delay, sido):
            state["status"] = "error"
            return
        _log(state, f"📍 지역 지정 크롤 — {name} (지연 {delay}s · {SAVE_EVERY_PAGES}쪽마다 중간저장)")
        ck = _load_ckpt()
        _do_region(state, session, idx, sido, name, delay, ck, "📍")
        state["status"] = "done"
        _log(state, "🏁 종료")
    except Exception as e:
        import traceback
        _log(state, "❌ 예외: " + str(e))
        _log(state, traceback.format_exc())
        state["status"] = "error"
    finally:
        if state["status"] not in ("done", "error"):
            state["status"] = "done"


def _run_queue(state, sido_list, delay):
    """작업 등록 큐 — 사용자가 선택한 시도들을 등록 순서대로 순차 크롤. 시도마다 엑셀 1개.
    각 지역은 체크포인트/재개(완료 지역은 건너뜀). 중간에 멈춰도 다음 실행 때 이어서."""
    try:
        os.makedirs(_EXPORT_DIR, exist_ok=True)
        # 선택 코드 → (idx, sido, name) 해석 (SIDO 정의 순 유지)
        items = []
        seen = set()
        for code in (sido_list or []):
            try:
                cv = int(code)
            except Exception:
                continue
            if cv in seen:
                continue
            seen.add(cv)
            m = [(i, s, n) for i, (s, n) in enumerate(SIDO, 1) if s == cv]
            if m:
                items.append(m[0])
        if not items:
            _log(state, "❌ 등록된 지역이 없습니다.")
            state["status"] = "error"
            return
        session = _session()
        if not _warm(state, session, delay, None):
            state["status"] = "error"
            return
        state["region_total"] = len(items)
        names = ", ".join(n for _, _, n in items)
        _log(state, f"🧾 작업 등록 {len(items)}개 지역 순차 크롤 (지연 {delay}s · {SAVE_EVERY_PAGES}쪽마다 중간저장)")
        _log(state, f"   순서: {names}")
        ck = _load_ckpt()
        for k, (idx, sido, name) in enumerate(items, 1):
            if state.get("cancel"):
                _log(state, "⏹ 사용자 중지")
                break
            res = _do_region(state, session, idx, sido, name, delay, ck, f"[{k}/{len(items)}]")
            if res == "cancelled":
                break
        state["status"] = "done"
        _log(state, f"🏁 작업 종료 — 완료 {state.get('region_done', 0)}/{len(items)} 지역")
    except Exception as e:
        import traceback
        _log(state, "❌ 예외: " + str(e))
        _log(state, traceback.format_exc())
        state["status"] = "error"
    finally:
        if state["status"] not in ("done", "error"):
            state["status"] = "done"


# ── 체크포인트 / 엑셀 저장·로드 ─────────────────────────────
def _load_ckpt() -> dict:
    try:
        with open(_CKPT_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_ckpt(c: dict):
    try:
        tmp = _CKPT_PATH + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(c, f, ensure_ascii=False)
        os.replace(tmp, _CKPT_PATH)
    except Exception:
        pass


def _region_file(idx: int, name: str) -> str:
    safe = re.sub(r'[\\/:*?"<>|]', "", name)
    return os.path.join(_EXPORT_DIR, f"한방_지역_{idx:02d}_{safe}.xlsx")


def _save_rows(rows, path) -> str:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    # 정렬: 주소(신주소) 순. 동률이면 상호.
    rows = sorted(rows, key=lambda r: (str(r.get("신주소") or ""), str(r.get("상호") or "")))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "한방"
    head_fill = PatternFill("solid", fgColor="D9E1F2")
    for c, name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, 2):
        for c, name in enumerate(COLUMNS, 1):
            ws.cell(row=r, column=c, value=row.get(name, ""))
    for c, w in enumerate([34, 16, 42], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    tmp = path + ".tmp"
    wb.save(tmp)
    # 원자적 교체 — 대상이 잠겨있으면(Excel로 열어둠 등) 재시도, 끝내 실패하면 예외(호출측이 잡아 계속 진행)
    last_err = None
    for attempt in range(4):
        try:
            os.replace(tmp, path)
            return path
        except PermissionError as e:   # WinError 5: 대상 파일 잠김
            last_err = e
            time.sleep(0.6)
    raise RuntimeError("파일 잠김으로 저장 실패(열려있는 Excel을 닫아주세요): " + os.path.basename(path) + " / " + str(last_err))


def _safe_save(state, rows, path) -> bool:
    """저장 시도. 실패(파일 잠김 등)해도 예외 안 내고 False 반환 — 크롤이 죽지 않게."""
    try:
        _save_rows(rows, path)
        return True
    except Exception as e:
        _log(state, f"  ⚠ 저장 실패({e}) — 데이터는 메모리 유지, 다음 저장 때 재시도하며 계속 진행")
        return False


def _load_xlsx_rows(path):
    import openpyxl
    try:
        ws = openpyxl.load_workbook(path).active
        headers = [c.value for c in ws[1]]
        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            out.append({headers[i]: ("" if i >= len(r) or r[i] is None else r[i]) for i in range(len(headers))})
        return out
    except Exception:
        return []


# ============================================================
# 공개 API (mj_extensions 라우터에서 호출)
# ============================================================
def start(start_page: int = 1, end_page: int = 1, delay: float = 1.0, mode: str = "single", sido=None, sido_list=None):
    # 동시 실행 방지 — 진행 중 작업(전국/지역/단일/큐) 있으면 거부 (체크포인트 파일 경쟁 방지)
    with _runs_lock:
        if any(s.get("status") == "running" for s in _runs.values()):
            raise RuntimeError("이미 다른 한방 크롤이 실행 중입니다. 먼저 중지하세요.")
    delay = max(0.3, float(delay))
    sp = max(1, int(start_page))
    ep = max(sp, int(end_page))
    run_id = uuid.uuid4().hex[:12]
    state = {
        "lines": [], "status": "running", "lock": threading.Lock(),
        "cancel": False, "count": 0, "cur_page": 0,
        "start_page": sp, "end_page": ep,
        "total_pages_site": 0, "file": None, "mode": mode,
        "cur_region": "", "region_done": 0, "region_total": 0, "region_files": [],
    }
    with _runs_lock:
        _runs[run_id] = state
    if mode == "regions":
        t = threading.Thread(target=_run_regions, args=(state, delay), daemon=True)
    elif mode == "queue":
        t = threading.Thread(target=_run_queue, args=(state, list(sido_list or []), delay), daemon=True)
    elif mode == "region_one":
        t = threading.Thread(target=_run_one_region, args=(state, int(sido or 1), delay), daemon=True)
    else:
        t = threading.Thread(target=_run, args=(state, sp, ep, delay), daemon=True)
    t.start()
    return run_id


def logs(run_id: str, offset: int = 0):
    state = _runs.get(run_id)
    if not state:
        return None
    with state["lock"]:
        lines = state["lines"][offset:]
    return {
        "lines": lines,
        "status": state["status"],
        "count": state["count"],
        "cur_page": state["cur_page"],
        "start_page": state["start_page"],
        "end_page": state["end_page"],
        "total_pages_site": state["total_pages_site"],
        "has_file": bool(state["file"]),
        "file_name": os.path.basename(state["file"]) if state["file"] else "",
        "mode": state.get("mode", "single"),
        "cur_region": state.get("cur_region", ""),
        "region_done": state.get("region_done", 0),
        "region_total": state.get("region_total", 0),
        "region_files": state.get("region_files", []),
    }


def stop(run_id: str) -> bool:
    state = _runs.get(run_id)
    if not state:
        return False
    state["cancel"] = True
    return True


def file_path(run_id: str):
    state = _runs.get(run_id)
    if state and state.get("file") and os.path.isfile(state["file"]):
        return state["file"]
    return None


# ── 지역 총건수 / 저장 폴더 목록 / 파일명 다운로드 ───────────────
def _fetch_simple(session, url):
    try:
        r = session.get(url, headers=_HEADERS, timeout=15)
        r.encoding = "utf-8"
        txt = r.text
        return txt, detect_block(txt)
    except Exception:
        return "", False


def region_info(sido_code):
    """지역 선택 시 총 페이지/추정 건수 + 이미 받은 진행도(체크포인트). 사이트 1요청."""
    try:
        match = [(i, s, n) for i, (s, n) in enumerate(SIDO, 1) if s == int(sido_code)]
        if not match:
            return {"ok": False, "error": "unknown sido"}
        idx, sido, name = match[0]
        html, blocked = _fetch_simple(_session(), _list_page_url(1, sido))
        if blocked:
            return {"ok": False, "error": "WAF 차단 — 잠시 후"}
        if not html:
            return {"ok": False, "error": "응답 없음"}
        total = _detect_total_pages(html)
        per = len(parse_list(html)) or 10
        ck = _load_ckpt()
        info = ck.get(str(idx), {})
        return {"ok": True, "name": name, "total_pages": total, "per_page": per,
                "est_count": total * per, "done": bool(info.get("done")),
                "last_page": int(info.get("last_page") or 0)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def regions_overview(refresh=False):
    """17개 시도 전체 목록 + 총페이지/추정건수(캐시) + 현재 진행도(체크포인트, 라이브).
    refresh=True면 사이트에서 총페이지 다시 조회(지역당 1요청, ~0.4s 페이싱)."""
    cache = {}
    if not refresh:
        try:
            cache = json.load(open(_TOTALS_CACHE, encoding="utf-8"))
        except Exception:
            cache = {}
    ck = _load_ckpt()
    session = None
    out = []
    fetched = False
    for idx, (sido, name) in enumerate(SIDO, 1):
        total = cache.get(str(sido))
        if not total:
            if session is None:
                session = _session()
            html, blocked = _fetch_simple(session, _list_page_url(1, sido))
            total = _detect_total_pages(html) if (html and not blocked) else 0
            cache[str(sido)] = total
            fetched = True
            time.sleep(0.4)   # WAF 회피 페이싱
        info = ck.get(str(idx), {})
        out.append({
            "idx": idx, "sido": sido, "name": name,
            "total_pages": int(total or 0), "est_count": int(total or 0) * 10,
            "done": bool(info.get("done")), "last_page": int(info.get("last_page") or 0),
            "started_at": info.get("started_at", ""), "done_at": info.get("done_at", ""),
        })
    if fetched:
        try:
            os.makedirs(_EXPORT_DIR, exist_ok=True)
            json.dump(cache, open(_TOTALS_CACHE, "w", encoding="utf-8"), ensure_ascii=False)
        except Exception:
            pass
    return {"ok": True, "regions": out}


def list_export_files():
    """저장 폴더(hanbang_exports)의 xlsx 목록 (이름/크기/건수/시각)."""
    out = []
    try:
        if not os.path.isdir(_EXPORT_DIR):
            return out
        import openpyxl
        for fn in sorted(os.listdir(_EXPORT_DIR)):
            if not fn.lower().endswith(".xlsx") or fn.endswith(".tmp"):
                continue
            fp = os.path.join(_EXPORT_DIR, fn)
            sz = 0; mt = ""; rows = None
            try:
                sz = os.path.getsize(fp)
                mt = time.strftime("%Y-%m-%d %H:%M", time.localtime(os.path.getmtime(fp)))
            except Exception:
                pass
            try:
                wb = openpyxl.load_workbook(fp, read_only=True)
                rows = wb.active.max_row - 1
                wb.close()
            except Exception:
                rows = None
            out.append({"name": fn, "size": sz, "mtime": mt, "rows": rows})
    except Exception:
        pass
    return out


def file_path_by_name(name):
    """저장 폴더 내 파일명으로 경로 반환 (경로 탈출 방지)."""
    if not name or "/" in name or "\\" in name or ".." in name:
        return None
    fp = os.path.join(_EXPORT_DIR, name)
    if os.path.isfile(fp) and fp.lower().endswith(".xlsx"):
        return fp
    return None
